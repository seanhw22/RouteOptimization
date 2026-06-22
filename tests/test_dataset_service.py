"""Smoke tests for datasets.services: validate_frames, save_dataset, parse_uploaded."""

import io
from pathlib import Path

import pandas as pd
import pytest
from django.conf import settings
from openpyxl import load_workbook

from datasets.services import DatasetValidationError, parse_uploaded, save_dataset, validate_frames


def test_dataset_template_has_required_blank_sheets():
    template_path = Path(settings.BASE_DIR) / 'static' / 'datasets' / 'dataset_template.xlsx'
    workbook = load_workbook(template_path, read_only=True)

    expected_headers = {
        'depots': ['depot_id', 'x', 'y'],
        'customers': ['customer_id', 'x', 'y', 'deadline_hours'],
        'vehicles': [
            'vehicle_id', 'depot_id', 'vehicle_type', 'capacity_kg',
            'max_operational_hrs', 'speed_kmh',
        ],
        'items': ['item_id', 'weight_kg', 'expiry_hours'],
        'orders': ['customer_id', 'item_id', 'quantity'],
    }

    assert workbook.sheetnames == list(expected_headers)
    for sheet_name, headers in expected_headers.items():
        sheet = workbook[sheet_name]
        assert [cell.value for cell in sheet[1]] == headers
        assert sheet.max_row == 1


# ── validate_frames ──────────────────────────────────────────────────────────

def test_validate_frames_rejects_missing_column(minimal_frames):
    bad = dict(minimal_frames)
    bad['depots'] = minimal_frames['depots'].drop(columns=['depot_id'])
    with pytest.raises(DatasetValidationError, match='depot_id'):
        validate_frames(bad)


def test_validate_frames_rejects_duplicate_pk(minimal_frames):
    bad = dict(minimal_frames)
    first_row = minimal_frames['depots'].iloc[[0]]
    bad['depots'] = pd.concat([minimal_frames['depots'], first_row], ignore_index=True)
    with pytest.raises(DatasetValidationError, match='D1'):
        validate_frames(bad)


def test_validate_frames_rejects_bad_referential_integrity(minimal_frames):
    bad = dict(minimal_frames)
    bad['orders'] = pd.DataFrame({
        'customer_id': ['UNKNOWN'],
        'item_id': ['I1'],
        'quantity': [1],
    })
    with pytest.raises(DatasetValidationError, match='customer_id'):
        validate_frames(bad)


def test_validate_frames_passes_valid(minimal_frames):
    validate_frames(minimal_frames)  # should not raise


# ── save_dataset ──────────────────────────────────────────────────────────────

@pytest.mark.django_db
def test_save_dataset_persists_correct_counts(minimal_frames, test_user):
    dataset = save_dataset(
        name='Count Test',
        user=test_user,
        session_key='',
        is_guest=False,
        frames=minimal_frames,
    )
    assert dataset.depots.count() == 2
    assert dataset.customers.count() == 3
    assert dataset.vehicles.count() == 2
    assert dataset.items.count() == 1
    assert dataset.orders.count() == 3


@pytest.mark.django_db
def test_save_dataset_guest_sets_expiry_and_token(minimal_frames):
    dataset = save_dataset(
        name='Guest Test',
        user=None,
        session_key='s-key',
        is_guest=True,
        frames=minimal_frames,
    )
    assert dataset.expires_at is not None
    assert dataset.share_token is not None


@pytest.mark.django_db
def test_save_dataset_authenticated_no_expiry(minimal_frames, test_user):
    dataset = save_dataset(
        name='Auth Test',
        user=test_user,
        session_key='',
        is_guest=False,
        frames=minimal_frames,
    )
    assert dataset.expires_at is None


# ── parse_uploaded ────────────────────────────────────────────────────────────

def _csv_file(df):
    buf = io.BytesIO()
    df.to_csv(buf, index=False)
    buf.seek(0)
    return buf


def test_parse_uploaded_five_csvs(minimal_frames):
    form_files = {f'{entity}_csv': _csv_file(df) for entity, df in minimal_frames.items()}
    result = parse_uploaded(form_files)
    assert set(result.keys()) == {'depots', 'customers', 'vehicles', 'items', 'orders'}
